from flask import Flask
from datetime import datetime
from openpyxl import Workbook, load_workbook
from flask import request, jsonify
from models.Database import Database
import os

database = Database()

api = Flask(__name__)

@api.route('/')
def welcome():
    return "Welcome to the CheckIN API"
    
@api.route('/checkin', methods=['POST'])
def checkin():
    return log_arrival(request) 
    
@api.route('/checkout', methods=['POST'])
def checkout():
    return log_departure(request)
    
def log_departure(request):
    """ 
        Logs the time and date of the person that 
        checked out
    """
    from openpyxl.util import cell as cell_util
    wb = load_wb() 
    ws = wb.active
    checkout_datetime = datetime.today() 
    
    target_col = cell_util.column_index_from_string('E')
    
    for cell in ws.iter_cols(min_col=target_col, max_col=target_col):
        if cell.value == request.json['id']: 
            cell.offset(column=1).value(checkout_datetime.time().isoformat());

    return jsonify(message="Check out confirmed", 
                   confirmed_time=checkout_datetime.time().isoformat())

def log_arrival(request):
    """
       Logs the time, date and IP address of the 
       person that checked in 
    """
    
    wb = load_wb()
    ws = wb.active
    checkin_datetime = datetime.today()
    
    last_row = find_document_end(ws)

    
    # First Column containing the IP address of 
    # the person checking in
    # TODO: change to include the GUID of the person that is
    ws['A'+str(last_row+1)].value = request.environ['REMOTE_ADDR']
    
    # Second Column containing the date the 
    # person logged in
    ws['B'+str(last_row+1)].value = checkin_datetime.date().isoformat()

    # Third Collumn containing the date the 
    # person logged in     
    ws['C'+str(last_row+1)].value = checkin_datetime.time().isoformat()
    
    # Fourth Column containing the location where
    # the person is checking in from
    ws['D'+str(last_row+1)].value = request.json['loc']
    
    # Fifth Column contains the Id of the person 
    # who logged in
    ws['E'+str(last_row+1)].value = request.json['id'];
    
    save_wb(wb)
    
    
    
    return jsonify(message="Check-In Logged", 
                   confirmed_time=checkin_dateime.time().isoformat(),
                   )

def create_wb():
    return Workbook()    

def save(wb, data):
    save_wb(wb)
    
def save_db(data):
    # Only saving log times on the API 
def save_wb(wb):
    wb.save("checkin_test.xlsx")
    
def find_document_end(ws):
    return ws.max_row
    
def load_wb():
    wb = None
    try:
        wb = load_workbook('checkin_test.xlsx')
    except IOError:
        # The file doesn't exist hence we have to create it 
        wb = create_wb()
    return wb
    
    
if __name__ == "__main__":
    api.run(os.getenv('IP', '0.0.0.0'), os.getenv('PORT', 8080))